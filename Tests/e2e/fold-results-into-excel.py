"""
Folds e2e/results.json (written by run.spec.js) back into
CSCS_Frontend_Test_Suite.xlsx -- updates the Outcome and Actual Result /
Notes columns for any test case that actually ran, leaves everything else
untouched.

Usage:
    pip install openpyxl --break-system-packages
    python3 fold-results-into-excel.py \
        --results results.json \
        --excel ../CSCS_Frontend_Test_Suite.xlsx
"""
import argparse
import json
import openpyxl


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--results", required=True)
    ap.add_argument("--excel", required=True)
    args = ap.parse_args()

    with open(args.results) as f:
        results = json.load(f)
    by_id = {r["id"]: r for r in results if r.get("status") != "Skipped"}

    wb = openpyxl.load_workbook(args.excel)
    ws = wb["Test Cases"]

    header = [c.value for c in ws[1]]
    id_col = header.index("Test Case ID") + 1
    outcome_col = header.index("Outcome") + 1
    notes_col = header.index("Actual Result / Notes") + 1

    updated = 0
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=id_col).value
        if tc_id in by_id:
            r = by_id[tc_id]
            status_map = {"Passed": "Passed", "Failed": "Failed"}
            ws.cell(row=row, column=outcome_col, value=status_map.get(r["status"], "Blocked"))
            note = r.get("note", "") or "Automated run — see e2e/screenshots/{}.png".format(tc_id)
            ws.cell(row=row, column=notes_col, value=note)
            updated += 1

    wb.save(args.excel)
    print(f"Updated {updated} test case(s) in {args.excel}")


if __name__ == "__main__":
    main()